#!/usr/bin/env python3
"""Recalculate a copy of an XLSX workbook with an isolated LibreOffice profile."""

from __future__ import annotations

import argparse
import json
import os
import shutil
import subprocess
import tempfile
import zipfile
from pathlib import Path

from openpyxl import load_workbook


EXCEL_ERRORS = {"#VALUE!", "#DIV/0!", "#REF!", "#NAME?", "#NULL!", "#NUM!", "#N/A"}


def formula_count(path: Path) -> int:
    workbook = load_workbook(path, read_only=True, data_only=False)
    try:
        return sum(
            1
            for sheet in workbook.worksheets
            for row in sheet.iter_rows()
            for cell in row
            if cell.data_type == "f"
        )
    finally:
        workbook.close()


def external_links(path: Path) -> bool:
    with zipfile.ZipFile(path) as package:
        names = set(package.namelist())
        return any(name.startswith("xl/externalLinks/") for name in names) or (
            "xl/connections.xml" in names
        )


def cached_errors(path: Path) -> list[str]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    found: list[str] = []
    try:
        for sheet in workbook.worksheets:
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.data_type == "e" or cell.value in EXCEL_ERRORS:
                        found.append(f"{sheet.title}!{cell.coordinate}:{cell.value}")
        return found
    finally:
        workbook.close()


def libreoffice_command(
    executable: str,
    source: Path,
    output_dir: Path,
    profile_dir: Path,
) -> list[str]:
    return [
        executable,
        f"-env:UserInstallation={profile_dir.resolve().as_uri()}",
        "--headless",
        "--nologo",
        "--nodefault",
        "--nolockcheck",
        "--norestore",
        "--convert-to",
        "xlsx",
        "--outdir",
        str(output_dir),
        str(source),
    ]


def recalculate(
    source: Path,
    output: Path,
    work_dir: Path,
    timeout: int = 60,
    replace: bool = False,
) -> dict[str, object]:
    source = source.resolve()
    output = output.resolve()
    work_dir = work_dir.resolve()

    if not source.is_file():
        return {"status": "error", "message": "输入文件不存在"}
    if source.suffix.lower() != ".xlsx" or output.suffix.lower() != ".xlsx":
        return {"status": "unsupported", "message": "本脚本只处理 .xlsx；不会转换 .xlsm"}
    if source == output:
        return {"status": "error", "message": "输出必须与输入不同，以保留原文件"}
    if output.exists() and not replace:
        return {"status": "error", "message": "输出已存在；确认后使用 --replace"}
    if not work_dir.is_dir():
        return {"status": "error", "message": "--work-dir 必须是已经建立的 workbench 目录"}

    try:
        formulas = formula_count(source)
    except Exception as exc:
        return {"status": "error", "message": f"无法读取工作簿：{exc}"}
    if formulas == 0:
        return {"status": "not_needed", "formula_count": 0, "message": "工作簿没有公式，未运行重新计算"}
    if external_links(source):
        return {"status": "unsupported", "message": "工作簿含外部链接；未自动更新或访问外部来源"}

    executable = shutil.which("soffice")
    if not executable:
        return {"status": "unavailable", "message": "未发现兼容的 soffice；未写入输出文件"}

    with tempfile.TemporaryDirectory(prefix="xlsx-recalc-", dir=work_dir) as temporary:
        temporary_path = Path(temporary)
        converted_dir = temporary_path / "converted"
        profile_dir = temporary_path / "profile"
        converted_dir.mkdir()
        profile_dir.mkdir()
        command = libreoffice_command(executable, source, converted_dir, profile_dir)
        environment = os.environ.copy()
        environment["SAL_USE_VCLPLUGIN"] = "svp"
        try:
            completed = subprocess.run(
                command,
                capture_output=True,
                text=True,
                timeout=timeout,
                env=environment,
                check=False,
            )
        except subprocess.TimeoutExpired:
            return {"status": "error", "message": "LibreOffice 重新计算超时；未写入输出文件"}
        if completed.returncode != 0:
            detail = (completed.stderr or completed.stdout or "未提供错误信息").strip()
            return {"status": "error", "message": f"LibreOffice 失败：{detail}"}

        converted = converted_dir / f"{source.stem}.xlsx"
        if not converted.is_file():
            return {"status": "error", "message": "LibreOffice 未生成预期工作簿"}
        try:
            errors = cached_errors(converted)
        except Exception as exc:
            return {"status": "error", "message": f"无法核对重新计算结果：{exc}"}
        if errors:
            return {
                "status": "errors_found",
                "formula_count": formulas,
                "error_count": len(errors),
                "errors": errors[:20],
                "message": "重新计算结果含公式错误；未写入正式输出",
            }

        output.parent.mkdir(parents=True, exist_ok=True)
        staging = output.with_name(f".{output.name}.recalc.tmp")
        try:
            shutil.copy2(converted, staging)
            os.replace(staging, output)
        finally:
            staging.unlink(missing_ok=True)
        return {
            "status": "success",
            "formula_count": formulas,
            "error_count": 0,
            "output": str(output),
        }


def main() -> int:
    parser = argparse.ArgumentParser(
        description="使用隔离的 LibreOffice 配置重新计算 XLSX 副本，不覆盖输入文件"
    )
    parser.add_argument("source", type=Path)
    parser.add_argument("output", type=Path)
    parser.add_argument("--work-dir", required=True, type=Path)
    parser.add_argument("--timeout", type=int, default=60)
    parser.add_argument("--replace", action="store_true")
    args = parser.parse_args()
    result = recalculate(
        args.source,
        args.output,
        args.work_dir,
        timeout=args.timeout,
        replace=args.replace,
    )
    print(json.dumps(result, ensure_ascii=False, indent=2))
    return 0 if result["status"] in {"success", "not_needed"} else 2


if __name__ == "__main__":
    raise SystemExit(main())
